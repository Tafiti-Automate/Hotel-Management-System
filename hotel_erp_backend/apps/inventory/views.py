import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction
from django.db.models import Exists, OuterRef, Q
from django.utils import timezone
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet

from apps.employees.models import Employee
from apps.inventory.models import (
    Category,
    DepartmentConsumption,
    InventoryBalance,
    InventoryBatch,
    ReorderRule,
    Item,
    ItemUnitPrice,
    StockAdjustment,
    StockAdjustmentItem,
    StockCount,
    StockCountItem,
    StockIssue,
    StockIssueItem,
    StockLedger,
    StockTransfer,
    StockTransferItem,
    StoreLocation,
    StoreKeeperAssignment,
    StoreRequisition,
    StoreRequisitionItem,
    StoreReturn,
    StoreReturnItem,
    SupplierItemPrice,
    UnitOfMeasure,
)
from apps.inventory.serializers import (
    CategorySerializer,
    DepartmentConsumptionSerializer,
    InventoryBalanceSerializer,
    InventoryBatchSerializer,
    ReorderRuleSerializer,
    ItemSerializer,
    ItemUnitPriceSerializer,
    StockAdjustmentItemSerializer,
    StockAdjustmentSerializer,
    StockCountItemSerializer,
    StockCountSerializer,
    StockIssueItemSerializer,
    StockIssueSerializer,
    StockLedgerSerializer,
    StockTransferItemSerializer,
    StockTransferSerializer,
    StoreLocationSerializer,
    StoreKeeperAssignmentSerializer,
    StoreRequisitionItemSerializer,
    StoreRequisitionSerializer,
    StoreReturnItemSerializer,
    StoreReturnSerializer,
    SupplierItemPriceSerializer,
    UnitOfMeasureSerializer,
)
from apps.procurement.models import RequisitionItem
from core.mixins.viewsets import CreatedByModelMixin
from core.constants.choices import ItemBusinessType, StoreRequisitionStatus


def raise_drf_validation_error(error):
    raise ValidationError(getattr(error, "messages", str(error)))


def enforce_readiness(readiness):
    if readiness["blockers"]:
        raise ValidationError(
            {
                "detail": "Complete the required steps before continuing.",
                "blockers": readiness["blockers"],
                "warnings": readiness["warnings"],
            }
        )


MASTER_DATA_ROLES = ("System Administrator", "Cost Controller")


def has_role(user, *roles):
    return bool(
        user
        and user.is_authenticated
        and (user.is_superuser or user.groups.filter(name__in=roles).exists())
    )


def is_department_head(user):
    return bool(
        user
        and user.is_authenticated
        and user.groups.filter(name="Department Head").exists()
    )


def assigned_store_ids(user):
    employee = getattr(user, "employee_profile", None)
    if not employee:
        return StoreLocation.objects.none().values_list("pk", flat=True)
    return StoreKeeperAssignment.objects.filter(
        employee=employee,
        is_active=True,
        store__is_active=True,
    ).values_list("store_id", flat=True)


def scope_store_requisitions(queryset, user):
    if user.is_superuser or has_role(user, "System Administrator"):
        return queryset
    employee = getattr(user, "employee_profile", None)
    if not employee:
        return queryset.none()
    if is_department_head(user):
        return queryset.filter(
            department=employee.department,
            requested_by__branch=employee.branch,
        ).exclude(status=StoreRequisitionStatus.DRAFT)
    if has_role(user, "Store Keeper"):
        # New requests already carry the requester's chosen issuing store. Keep
        # the branch fallback only for legacy submitted requests with no store.
        store_ids = assigned_store_ids(user)
        branch_ids = StoreLocation.objects.filter(pk__in=store_ids).values_list("branch_id", flat=True)
        return queryset.filter(department_approved_by__isnull=False).filter(
            Q(store_id__in=store_ids)
            | Q(
                status=StoreRequisitionStatus.SUBMITTED,
                store__isnull=True,
                requested_by__branch_id__in=branch_ids,
            )
        ).distinct()
    # Only an explicitly assigned Requester may originate/track Department requests.
    if has_role(user, "Requester"):
        return queryset.filter(requested_by=employee)
    # All other operational roles start from their own workflow document/stage.
    return queryset.none()


class CostControllerAuthorityMixin:
    """Cost Controller owns item, UOM, conversion and supplier quote master data."""

    def _require_master_data_authority(self):
        if not has_role(self.request.user, *MASTER_DATA_ROLES):
            raise PermissionDenied("Only the Cost Controller can maintain procurement master data.")

    def perform_create(self, serializer):
        self._require_master_data_authority()
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._require_master_data_authority()
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._require_master_data_authority()
        super().perform_destroy(instance)


class CategoryViewSet(CostControllerAuthorityMixin, CreatedByModelMixin, ModelViewSet):
    queryset = Category.objects.select_related("parent")
    serializer_class = CategorySerializer
    filterset_fields = ("parent", "is_active")
    search_fields = ("name", "code", "description", "parent__name")
    ordering_fields = ("name", "code", "parent__name", "created_at")


class UnitOfMeasureViewSet(CostControllerAuthorityMixin, CreatedByModelMixin, ModelViewSet):
    queryset = UnitOfMeasure.objects.all()
    serializer_class = UnitOfMeasureSerializer
    filterset_fields = ("is_active",)
    search_fields = ("name", "abbreviation")
    ordering_fields = ("name", "created_at")


class ItemViewSet(CostControllerAuthorityMixin, CreatedByModelMixin, ModelViewSet):
    queryset = Item.objects.select_related("category", "base_unit").annotate(
        _has_unit_price_usage=Exists(
            ItemUnitPrice.objects.filter(item_id=OuterRef("pk"))
        ),
        _has_inventory_balance_usage=Exists(
            InventoryBalance.objects.filter(item_id=OuterRef("pk"))
        ),
        _has_purchase_requisition_usage=Exists(
            RequisitionItem.objects.filter(item_id=OuterRef("pk"))
        ),
        _has_store_requisition_usage=Exists(
            StoreRequisitionItem.objects.filter(item_id=OuterRef("pk"))
        ),
    )
    serializer_class = ItemSerializer
    filterset_fields = ("category", "unit", "base_unit", "is_active")
    search_fields = ("name", "sku", "barcode", "brand", "category__name")
    ordering_fields = ("name", "sku", "reorder_level", "created_at")

    @action(detail=False, methods=["post"], url_path="import")
    def import_items(self, request):
        """Atomically create or update Articles from a CSV or Excel workbook."""
        self._require_master_data_authority()
        upload = request.FILES.get("file")
        if not upload:
            raise ValidationError({"file": "Choose a CSV or Excel (.xlsx) item file."})
        try:
            rows = self._spreadsheet_rows(upload)
        except (ValueError, UnicodeDecodeError) as error:
            raise ValidationError({"file": str(error)})

        created = updated = 0
        errors = []
        with transaction.atomic():
            for number, raw in enumerate(rows, start=2):
                try:
                    result = self._import_item_row(raw, request)
                    created += result == "created"
                    updated += result == "updated"
                except Exception as error:
                    detail = (
                        getattr(error, "detail", None)
                        or getattr(error, "message_dict", None)
                        or str(error)
                    )
                    errors.append({"row": number, "error": detail})
            if errors:
                transaction.set_rollback(True)
        if errors:
            return Response(
                {
                    "detail": "Nothing was imported. Correct the listed rows and try again.",
                    "errors": errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response({"created": created, "updated": updated, "total": created + updated})

    @staticmethod
    def _spreadsheet_rows(upload):
        name = upload.name.lower()
        if name.endswith(".csv"):
            content = upload.read().decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(content)))
        if name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as error:
                raise ValueError(
                    "Excel support is not installed on the server; upload CSV instead."
                ) from error
            sheet = load_workbook(upload, read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(values)]
            except StopIteration:
                return []
            return [
                dict(zip(headers, row))
                for row in values
                if any(value is not None for value in row)
            ]
        raise ValueError("Unsupported file type. Upload .csv or .xlsx.")

    @staticmethod
    def _import_item_row(raw, request):
        normalized = {
            str(key).strip().lower().replace(" ", "_"): value
            for key, value in raw.items()
        }
        major_name = str(normalized.get("major_group") or "").strip()
        group_name = str(normalized.get("item_group") or normalized.get("category") or "").strip()
        item_name = str(normalized.get("item_name") or normalized.get("name") or "").strip()
        sku = str(normalized.get("sku") or normalized.get("sku_code") or "").strip().upper()
        unit_name = str(normalized.get("base_unit") or normalized.get("unit") or "").strip()
        if not all((major_name, group_name, item_name, sku, unit_name)):
            raise ValueError(
                "major_group, item_group, item_name, sku and base_unit are required"
            )

        major = Category.objects.filter(name__iexact=major_name).first()
        if major and major.parent_id:
            raise ValueError(f"'{major_name}' is an item group, not a major group")
        if not major:
            major = Category.objects.create(name=major_name, created_by=request.user)

        group = Category.objects.filter(name__iexact=group_name).first()
        if group and group.parent_id != major.id:
            raise ValueError(f"Item group '{group_name}' does not belong to '{major_name}'")
        if group and not group.parent_id:
            raise ValueError(f"'{group_name}' is a major group, not an item group")
        if not group:
            group = Category.objects.create(
                name=group_name,
                parent=major,
                created_by=request.user,
            )

        unit = (
            UnitOfMeasure.objects.filter(abbreviation__iexact=unit_name).first()
            or UnitOfMeasure.objects.filter(name__iexact=unit_name).first()
        )
        if not unit:
            raise ValueError(f"Unknown base unit '{unit_name}'")

        business_value = str(
            normalized.get("business_type")
            or normalized.get("business_classification")
            or ItemBusinessType.CONSUMABLE_EXPENSE
        ).strip()
        business_lookup = {
            str(label).lower(): value for value, label in ItemBusinessType.choices
        }
        business_type = business_lookup.get(
            business_value.lower().replace("_", " "),
            business_value.lower().replace(" ", "_"),
        )
        if business_type not in ItemBusinessType.values:
            raise ValueError(f"Unknown business type '{business_value}'")

        def decimal_value(key, default="0.00", nullable=False):
            value = normalized.get(key)
            if value in (None, ""):
                return None if nullable else Decimal(default)
            try:
                return Decimal(str(value))
            except (InvalidOperation, TypeError, ValueError) as error:
                raise ValueError(f"{key} must be a number") from error

        def boolean_value(key, default=False):
            value = normalized.get(key)
            if value in (None, ""):
                return default
            return str(value).strip().lower() not in {"no", "false", "0", "inactive"}

        payload = {
            "category": group.pk,
            "name": item_name,
            "sku": sku,
            "brand": str(normalized.get("brand") or "").strip(),
            "description": str(normalized.get("description") or "").strip(),
            "barcode": str(normalized.get("barcode") or "").strip() or None,
            "unit": unit.abbreviation or unit.name,
            "base_unit": unit.pk,
            "reorder_level": decimal_value("reorder_level"),
            "maximum_level": decimal_value("maximum_level", nullable=True),
            "batch_tracking": boolean_value("batch_tracking"),
            "expiry_tracking": boolean_value("expiry_tracking"),
            "business_type": business_type,
            "is_active": boolean_value("is_active", True),
        }
        existing = Item.objects.filter(sku__iexact=sku).first()
        serializer = ItemSerializer(
            existing,
            data=payload,
            context={"request": request, "allow_sku_input": True},
        ) if existing else ItemSerializer(
            data=payload,
            context={"request": request, "allow_sku_input": True},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(
            created_by=existing.created_by if existing else request.user,
        )
        return "updated" if existing else "created"


class ItemUnitPriceViewSet(CostControllerAuthorityMixin, CreatedByModelMixin, ModelViewSet):
    queryset = ItemUnitPrice.objects.select_related("item", "unit")
    serializer_class = ItemUnitPriceSerializer
    filterset_fields = ("item", "unit", "role", "is_active")
    search_fields = ("item__name", "item__sku", "unit__name")
    ordering_fields = ("conversion_factor", "selling_price", "created_at")

    def perform_destroy(self, instance):
        self._require_master_data_authority()
        if instance.is_used_in_transactions():
            raise ValidationError(
                "This conversion is already used by a transaction and cannot be deleted."
            )
        super().perform_destroy(instance)


class StoreLocationViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreLocation.objects.select_related("branch")
    serializer_class = StoreLocationSerializer
    filterset_fields = ("branch", "is_active", "is_default")
    search_fields = ("name", "address", "branch__name")
    ordering_fields = ("name", "created_at")

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_role(self.request.user, "Store Keeper") and not has_role(
            self.request.user, "System Administrator", "General Manager"
        ):
            return queryset.filter(pk__in=assigned_store_ids(self.request.user))
        return queryset


class StoreKeeperAssignmentViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreKeeperAssignment.objects.select_related("store", "employee", "employee__user")
    serializer_class = StoreKeeperAssignmentSerializer
    filterset_fields = ("store", "employee", "is_active")
    search_fields = ("store__name", "employee__user__username", "employee__user__employee_code")
    ordering_fields = ("store__name", "employee__user__username", "created_at")

    def _require_administrator(self):
        if not has_role(self.request.user, "System Administrator"):
            raise PermissionDenied("Only the System Administrator can assign Store Keepers to stores.")

    def perform_create(self, serializer):
        self._require_administrator()
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._require_administrator()
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._require_administrator()
        super().perform_destroy(instance)


class InventoryBalanceViewSet(ReadOnlyModelViewSet):
    queryset = InventoryBalance.objects.select_related("item", "store")
    serializer_class = InventoryBalanceSerializer
    filterset_fields = ("item", "store")
    search_fields = ("item__name", "item__sku", "store__name")
    ordering_fields = ("quantity_in_stock", "reorder_level", "last_updated")

    def get_queryset(self):
        queryset = super().get_queryset()
        if has_role(self.request.user, "Store Keeper") and not has_role(
            self.request.user, "System Administrator", "General Manager"
        ):
            return queryset.filter(store_id__in=assigned_store_ids(self.request.user))
        return queryset

    @action(detail=True, methods=["post"], url_path="reconcile-reservation")
    def reconcile_reservation(self, request, pk=None):
        if not (
            request.user.is_superuser
            or request.user.groups.filter(name__in=("System Administrator", "Store Keeper")).exists()
        ):
            raise PermissionDenied("Only the Store Keeper can reconcile reserved stock.")
        with transaction.atomic():
            balance = InventoryBalance.objects.select_for_update().get(pk=self.get_object().pk)
            lines = StoreRequisitionItem.objects.filter(
                item=balance.item,
                requisition__store=balance.store,
                requisition__status__in=(
                    StoreRequisitionStatus.APPROVED,
                    StoreRequisitionStatus.PARTIALLY_APPROVED,
                    StoreRequisitionStatus.PARTIALLY_ISSUED,
                ),
            )
            calculated = sum(
                (line.outstanding_quantity for line in lines if line.outstanding_quantity > 0),
                Decimal("0.00"),
            )
            balance.quantity_reserved = calculated
            balance.save(update_fields=["quantity_reserved", "updated_at"])
        return Response(self.get_serializer(balance).data)


class SupplierItemPriceViewSet(CostControllerAuthorityMixin, CreatedByModelMixin, ModelViewSet):
    queryset = SupplierItemPrice.objects.select_related("supplier", "item", "item__category", "unit")
    serializer_class = SupplierItemPriceSerializer
    filterset_fields = ("supplier", "item", "item__category", "unit", "is_preferred", "is_active")
    search_fields = ("supplier__name", "item__name", "item__sku", "supplier_sku")
    ordering_fields = ("unit_price", "lead_time_days", "minimum_order_quantity", "last_quoted_at", "created_at")

    @action(detail=True, methods=["get"])
    def history(self, request, pk=None):
        from apps.inventory.serializers import SupplierItemPriceHistorySerializer

        price = self.get_object()
        return Response(SupplierItemPriceHistorySerializer(price.price_history.all(), many=True).data)

    @action(detail=False, methods=["post"], url_path="import")
    def import_catalogue(self, request):
        self._require_master_data_authority()
        upload = request.FILES.get("file")
        if not upload:
            raise ValidationError({"file": "Choose a CSV or Excel (.xlsx) supplier catalogue."})
        try:
            rows = self._spreadsheet_rows(upload)
        except (ValueError, UnicodeDecodeError) as error:
            raise ValidationError({"file": str(error)})

        created = updated = 0
        errors = []
        with transaction.atomic():
            for number, raw in enumerate(rows, start=2):
                try:
                    result = self._import_row(raw, request)
                    created += result == "created"
                    updated += result == "updated"
                except Exception as error:  # collect row-specific validation feedback
                    detail = getattr(error, "detail", None) or getattr(error, "message_dict", None) or str(error)
                    errors.append({"row": number, "error": detail})
            if errors:
                transaction.set_rollback(True)
        if errors:
            return Response(
                {"detail": "Nothing was imported. Correct the listed rows and try again.", "errors": errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return Response({"created": created, "updated": updated, "total": created + updated})

    @staticmethod
    def _spreadsheet_rows(upload):
        name = upload.name.lower()
        if name.endswith(".csv"):
            content = upload.read().decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(content)))
        if name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as error:
                raise ValueError("Excel support is not installed on the server; upload CSV instead.") from error
            sheet = load_workbook(upload, read_only=True, data_only=True).active
            values = sheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(values)]
            except StopIteration:
                return []
            return [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]
        raise ValueError("Unsupported file type. Upload .csv or .xlsx.")

    @staticmethod
    def _import_row(raw, request):
        from apps.vendors.models import Supplier

        normalized = {str(key).strip().lower().replace(" ", "_"): value for key, value in raw.items()}
        supplier_code = str(normalized.get("supplier_code") or "").strip()
        supplier_name = str(normalized.get("supplier") or normalized.get("supplier_name") or "").strip()
        item_sku = str(normalized.get("item_sku") or normalized.get("sku") or "").strip()
        unit_name = str(normalized.get("unit") or normalized.get("purchase_unit") or "").strip()
        if not item_sku or not (supplier_code or supplier_name):
            raise ValueError("supplier/supplier_code and item_sku are required")
        supplier_query = Supplier.objects.filter(supplier_code__iexact=supplier_code) if supplier_code else Supplier.objects.filter(name__iexact=supplier_name)
        supplier = supplier_query.get()
        item = Item.objects.get(sku__iexact=item_sku)
        unit = UnitOfMeasure.objects.filter(abbreviation__iexact=unit_name).first() or UnitOfMeasure.objects.filter(name__iexact=unit_name).first()
        if unit_name and not unit:
            raise ValueError(f"Unknown purchase unit '{unit_name}'")
        try:
            price = Decimal(str(normalized.get("unit_price") or normalized.get("price") or ""))
        except InvalidOperation as error:
            raise ValueError("unit_price must be a number") from error
        effective = normalized.get("effective_from") or date.today()
        if isinstance(effective, datetime):
            effective = effective.date()
        payload = {
            "supplier": supplier.pk,
            "item": item.pk,
            "unit": unit.pk if unit else None,
            "supplier_sku": str(normalized.get("supplier_sku") or "").strip(),
            "unit_price": price,
            "currency": str(normalized.get("currency") or "UGX").strip().upper(),
            "effective_from": effective,
            "minimum_order_quantity": normalized.get("minimum_order_quantity") or 1,
            "lead_time_days": normalized.get("lead_time_days") or 0,
            "quotation_reference": str(normalized.get("quotation_reference") or "").strip(),
            "quotation_valid_until": normalized.get("quotation_valid_until") or None,
            # Supplier quotations are alternatives for Procurement comparison.
            # There is no single preferred-supplier rule in the client workflow.
            "is_preferred": False,
            "is_active": str(normalized.get("is_active") or "yes").strip().lower() not in {"no", "false", "0", "inactive"},
        }
        existing = SupplierItemPrice.objects.filter(supplier=supplier, item=item).first()
        serializer = SupplierItemPriceSerializer(existing, data=payload, context={"request": request}) if existing else SupplierItemPriceSerializer(data=payload, context={"request": request})
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=request.user if not existing else existing.created_by)
        if existing and existing.price_history.exists():
            existing.price_history.first().source = "import"
            existing.price_history.first().save(update_fields=["source", "updated_at"])
        return "updated" if existing else "created"


class StockLedgerViewSet(ReadOnlyModelViewSet):
    queryset = StockLedger.objects.select_related("item", "store")
    serializer_class = StockLedgerSerializer
    filterset_fields = ("item", "store", "reference_type", "reference_id")
    search_fields = ("item__name", "item__sku", "store__name", "reference_id")
    ordering_fields = ("created_at", "quantity_in", "quantity_out")


class InventoryBatchViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = InventoryBatch.objects.select_related("item", "store", "purchase_order_item")
    serializer_class = InventoryBatchSerializer
    filterset_fields = ("item", "store", "expiry_date", "purchase_order_item")
    search_fields = ("item__name", "item__sku", "store__name")
    ordering_fields = ("expiry_date", "remaining_quantity", "unit_cost", "created_at")


class StockTransferViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockTransfer.objects.select_related(
        "from_store",
        "to_store",
        "requested_by",
        "approved_by",
    )
    serializer_class = StockTransferSerializer
    filterset_fields = ("from_store", "to_store", "status", "requested_by", "approved_by")
    search_fields = ("from_store__name", "to_store__name", "note")
    ordering_fields = ("status", "required_date", "created_at")

    @action(detail=True, methods=["post"], url_path="assign-store")
    def assign_store(self, request, pk=None):
        requisition = self.get_object()
        if not has_role(request.user, "System Administrator", "Store Keeper"):
            raise PermissionDenied("Only the Store Keeper can choose the destination store.")
        if requisition.status != StoreRequisitionStatus.SUBMITTED:
            raise ValidationError("The destination store can only be selected for a submitted Department request.")
        store_id = request.data.get("store")
        if not store_id:
            raise ValidationError({"store": "Select the destination store."})
        store = StoreLocation.objects.filter(pk=store_id, is_active=True).first()
        if not store:
            raise ValidationError({"store": "The selected destination store is not available."})
        if not has_role(request.user, "System Administrator") and not StoreKeeperAssignment.objects.filter(
            store=store, employee=getattr(request.user, "employee_profile", None), is_active=True
        ).exists():
            raise PermissionDenied("You can only route requests to a store assigned to you.")
        requisition.store = store
        requisition.save(update_fields=["store", "updated_at"])
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer.approve(approved_by=getattr(request.user, "employee_profile", None))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=["post"], url_path="dispatch")
    def dispatch_transfer(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer.dispatch(dispatched_by=getattr(request.user, "employee_profile", None))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=["post"])
    def receive(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer.receive(received_by=getattr(request.user, "employee_profile", None))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(transfer).data)

    @action(detail=True, methods=["post"])
    def apply(self, request, pk=None):
        transfer = self.get_object()
        try:
            transfer.apply_inventory_changes()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        serializer = self.get_serializer(transfer)
        return Response(serializer.data)


class StockTransferItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockTransferItem.objects.select_related("stock_transfer", "item", "unit")
    serializer_class = StockTransferItemSerializer
    filterset_fields = ("stock_transfer", "item", "unit")
    search_fields = ("item__name", "item__sku")
    ordering_fields = ("quantity", "base_quantity", "created_at")


class StockAdjustmentViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockAdjustment.objects.select_related("store", "approved_by")
    serializer_class = StockAdjustmentSerializer
    filterset_fields = ("store", "status", "approved_by")
    search_fields = ("reference", "reason", "note", "store__name")
    ordering_fields = ("status", "created_at")

    @action(detail=True, methods=["post"])
    def apply(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment.apply()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        serializer = self.get_serializer(adjustment)
        return Response(serializer.data)


class StockAdjustmentItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockAdjustmentItem.objects.select_related("stock_adjustment", "item", "unit")
    serializer_class = StockAdjustmentItemSerializer
    filterset_fields = ("stock_adjustment", "item", "unit")
    search_fields = ("item__name", "item__sku", "reason")
    ordering_fields = ("quantity_change", "created_at")

class ReorderRuleViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = ReorderRule.objects.select_related("item", "store", "preferred_supplier")
    serializer_class = ReorderRuleSerializer
    filterset_fields = ("item", "store", "preferred_supplier", "is_active")
    search_fields = ("item__name", "item__sku", "store__name", "preferred_supplier__name")
    ordering_fields = ("minimum_level", "reorder_quantity", "created_at")

    @action(detail=True, methods=["post"], url_path="create-purchase-requisition")
    def create_purchase_requisition(self, request, pk=None):
        reorder_rule = self.get_object()
        requester = getattr(request.user, "employee_profile", None)
        department = requester.department if requester else None
        try:
            purchase_requisition = reorder_rule.create_purchase_requisition(
                requester=requester,
                department=department,
                reason=request.data.get("reason", ""),
                created_by=request.user if request.user.is_authenticated else None,
            )
        except DjangoValidationError as error:
            raise_drf_validation_error(error)

        from apps.procurement.serializers import PurchaseRequisitionSerializer

        serializer = PurchaseRequisitionSerializer(
            purchase_requisition,
            context=self.get_serializer_context(),
        )
        return Response(serializer.data)


class StoreRequisitionViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreRequisition.objects.select_related(
        "department",
        "store",
        "requested_by",
        "approved_by",
        "department_approved_by",
        "procurement_requisition",
    )
    serializer_class = StoreRequisitionSerializer
    filterset_fields = ("department", "store", "requested_by", "approved_by", "status")
    search_fields = ("requisition_no", "purpose", "department__name", "store__name")
    ordering_fields = ("requisition_no", "status", "required_date", "created_at")
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return scope_store_requisitions(super().get_queryset(), self.request.user)

    def perform_create(self, serializer):
        if not has_role(self.request.user, "System Administrator", "Requester"):
            raise PermissionDenied("Only an assigned Requester can create a Department request.")
        super().perform_create(serializer)

    def _enforce_requester_edit(self, requisition):
        if has_role(self.request.user, "System Administrator"):
            return
        employee = getattr(self.request.user, "employee_profile", None)
        if (
            has_role(self.request.user, "Requester")
            and employee
            and requisition.requested_by_id == employee.id
            and requisition.status in (
                StoreRequisitionStatus.DRAFT,
                StoreRequisitionStatus.REJECTED,
            )
        ):
            return
        raise PermissionDenied("Only the requester can edit their draft store request.")

    def perform_update(self, serializer):
        self._enforce_requester_edit(serializer.instance)
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._enforce_requester_edit(instance)
        instance.delete()

    def _enforce_store_assignment(self, requisition):
        user = self.request.user
        if has_role(user, "System Administrator"):
            return
        if user.groups.filter(name="Store Keeper").exists() and StoreKeeperAssignment.objects.filter(
            store=requisition.store,
            employee=getattr(user, "employee_profile", None),
            is_active=True,
        ).exists():
            return
        raise PermissionDenied("Only the Store Keeper assigned to this store can process this request.")

    def _enforce_department_head(self, requisition):
        employee = getattr(self.request.user, "employee_profile", None)
        if not is_department_head(self.request.user) or not employee:
            raise PermissionDenied("Only the Department Head can decide this request.")
        if employee.department_id != requisition.department_id:
            raise PermissionDenied("You can only decide requests from your department.")
        requester_branch_id = getattr(requisition.requested_by, "branch_id", None)
        if requester_branch_id and employee.branch_id != requester_branch_id:
            raise PermissionDenied("You can only decide requests from your branch.")
        if employee.pk == requisition.requested_by_id:
            raise PermissionDenied("You cannot approve or reject your own request.")

    @action(detail=False, methods=["get"], url_path="store-options")
    def store_options(self, request):
        queryset = StoreLocation.objects.select_related("branch").filter(is_active=True)
        if not has_role(request.user, "System Administrator") and not request.user.is_superuser:
            employee = getattr(request.user, "employee_profile", None)
            if not employee or not employee.branch_id:
                queryset = queryset.none()
            elif has_role(request.user, "Store Keeper"):
                queryset = queryset.filter(pk__in=assigned_store_ids(request.user))
            else:
                queryset = queryset.filter(branch_id=employee.branch_id)
        return Response(StoreLocationSerializer(queryset.order_by("name"), many=True).data)

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        requisition = self.get_object()
        self._enforce_requester_edit(requisition)
        try:
            requisition.submit(actor=request.user)
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"], url_path="department-approve")
    def department_approve(self, request, pk=None):
        requisition = self.get_object()
        self._enforce_department_head(requisition)
        try:
            requisition.approve_department(
                approved_by=request.user.employee_profile,
                comments=request.data.get("comments", ""),
                item_quantities=request.data.get("items"),
            )
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        except Exception as error:
            # Record the full traceback in server logs and return a JSON error
            # with a request id.  This avoids the unhelpful generic HTML 500 that
            # previously reached the frontend and makes any remaining production-
            # database issue traceable without exposing database details.
            import logging
            request_id = getattr(request, "request_id", "")
            logging.getLogger(__name__).exception(
                "Unexpected HOD approval failure requisition=%s user=%s request_id=%s",
                requisition.pk,
                request.user.pk,
                request_id,
            )
            return Response(
                {
                    "detail": (
                        f"Department approval could not be saved ({type(error).__name__}). "
                        f"Reference: {request_id or requisition.pk}."
                    ),
                    "code": "hod_approval_failed",
                    "request_id": request_id,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # The frontend refreshes its queue after a successful action.  Returning
        # the entire requisition here is unnecessary and used to make approval
        # success depend on a second serialization/query path.
        return Response({
            "id": str(requisition.pk),
            "requisition_no": requisition.requisition_no,
            "status": StoreRequisitionStatus.SUBMITTED,
            "department_approved_at": requisition.department_approved_at,
            "department_approved_by": str(request.user.employee_profile.pk),
        })

    @action(detail=True, methods=["post"], url_path="assign-store")
    def assign_store(self, request, pk=None):
        requisition = self.get_object()
        if not has_role(request.user, "System Administrator", "Store Keeper"):
            raise PermissionDenied("Only the Store Keeper can change the issuing store after submission.")
        if requisition.status != StoreRequisitionStatus.SUBMITTED:
            raise ValidationError("The issuing store can only be changed for a submitted Department request.")
        store_id = request.data.get("store")
        if not store_id:
            raise ValidationError({"store": "Select the issuing store."})
        store = StoreLocation.objects.filter(pk=store_id, is_active=True).first()
        if not store:
            raise ValidationError({"store": "The selected issuing store is not available."})
        if not has_role(request.user, "System Administrator") and not StoreKeeperAssignment.objects.filter(
            store=store, employee=getattr(request.user, "employee_profile", None), is_active=True
        ).exists():
            raise PermissionDenied("You can only route requests to a store assigned to you.")
        requisition.store = store
        requisition.save(update_fields=["store", "updated_at"])
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        requisition = self.get_object()
        self._enforce_store_assignment(requisition)
        try:
            requisition.approve(
                approved_by=getattr(request.user, "employee_profile", None),
                comments=request.data.get("comments", ""),
            )
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"], url_path="send-to-procurement")
    def send_to_procurement(self, request, pk=None):
        requisition = self.get_object()
        self._enforce_store_assignment(requisition)
        try:
            purchase = requisition.create_procurement_requisition(
                created_by=request.user,
                reason=request.data.get("reason", ""),
            )
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        from apps.procurement.serializers import PurchaseRequisitionSerializer
        return Response(PurchaseRequisitionSerializer(purchase, context=self.get_serializer_context()).data)

    @action(detail=True, methods=["post"], url_path="resume-after-procurement")
    def resume_after_procurement(self, request, pk=None):
        requisition = self.get_object()
        self._enforce_store_assignment(requisition)
        try:
            requisition.resume_after_procurement(actor=request.user)
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        requisition = self.get_object()
        if requisition.status == StoreRequisitionStatus.PENDING_DEPARTMENT_APPROVAL:
            self._enforce_department_head(requisition)
        elif requisition.status == StoreRequisitionStatus.SUBMITTED and not requisition.store_id and has_role(request.user, "Store Keeper"):
            employee = getattr(request.user, "employee_profile", None)
            requester_branch_id = getattr(requisition.requested_by, "branch_id", None)
            if not employee or not StoreKeeperAssignment.objects.filter(
                employee=employee, is_active=True, store__branch_id=requester_branch_id
            ).exists():
                raise PermissionDenied("Only a Store Keeper assigned within this branch can reject the requisition.")
        else:
            self._enforce_store_assignment(requisition)
        try:
            requisition.reject(
                reason=request.data.get("reason", ""),
                actor=request.user,
            )
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        requisition = self.get_object()
        employee = getattr(request.user, "employee_profile", None)
        try:
            self._enforce_store_assignment(requisition)
            stores_control = True
        except PermissionDenied:
            stores_control = False
        requester_can_cancel = (
            employee
            and requisition.requested_by_id == employee.id
            and requisition.status in (
                StoreRequisitionStatus.DRAFT,
                StoreRequisitionStatus.PENDING_DEPARTMENT_APPROVAL,
                StoreRequisitionStatus.SUBMITTED,
            )
        )
        if not (stores_control or requester_can_cancel):
            raise PermissionDenied(
                "Only the requester may cancel an unapproved request; approved reservations require the Store Keeper."
            )
        try:
            requisition.cancel(actor=request.user)
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(requisition).data)


class StoreRequisitionItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreRequisitionItem.objects.select_related("requisition", "item", "unit")
    serializer_class = StoreRequisitionItemSerializer
    filterset_fields = ("requisition", "item", "unit")
    search_fields = ("requisition__requisition_no", "item__name", "item__sku")
    ordering_fields = ("quantity_requested", "quantity_approved", "quantity_issued", "created_at")
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        requisitions = scope_store_requisitions(
            StoreRequisition.objects.all(),
            self.request.user,
        )
        return super().get_queryset().filter(requisition__in=requisitions)

    def _enforce_line_edit(self, requisition):
        if has_role(self.request.user, "System Administrator"):
            return
        employee = getattr(self.request.user, "employee_profile", None)
        requester_edit = (
            has_role(self.request.user, "Requester")
            and employee
            and requisition.requested_by_id == employee.id
            and requisition.status in (
                StoreRequisitionStatus.DRAFT,
                StoreRequisitionStatus.REJECTED,
            )
        )
        keeper_edit = (
            has_role(self.request.user, "Store Keeper")
            and requisition.status == StoreRequisitionStatus.SUBMITTED
            and StoreKeeperAssignment.objects.filter(
                store=requisition.store,
                employee=employee,
                is_active=True,
            ).exists()
        )
        if requester_edit or keeper_edit:
            return
        raise PermissionDenied("You cannot change lines on this store request.")

    def perform_create(self, serializer):
        self._enforce_line_edit(serializer.validated_data["requisition"])
        super().perform_create(serializer)

    def perform_update(self, serializer):
        self._enforce_line_edit(serializer.instance.requisition)
        super().perform_update(serializer)

    @action(detail=True, methods=["post"], url_path="reject-line")
    def reject_line(self, request, pk=None):
        line = self.get_object()
        self._enforce_line_edit(line.requisition)
        if line.requisition.status != StoreRequisitionStatus.SUBMITTED:
            raise ValidationError("Store Keeper item rejection is only available while the request is awaiting Stores review.")
        if line.department_approved_limit <= Decimal("0.00"):
            raise ValidationError("This item was already rejected by the Department Head.")
        reason = str(request.data.get("reason") or "").strip()
        if not reason:
            raise ValidationError({"reason": "Enter the reason for rejecting this item."})
        with transaction.atomic():
            request_lines = list(
                StoreRequisitionItem.objects.select_for_update()
                .filter(requisition=line.requisition)
                .order_by("pk")
            )
            other_viable_lines = [
                candidate for candidate in request_lines
                if candidate.pk != line.pk
                and candidate.department_approved_limit > Decimal("0.00")
                and candidate.rejection_stage != "Store Keeper"
            ]
            if not other_viable_lines:
                raise ValidationError(
                    "This is the last remaining item. Use Reject entire requisition instead."
                )
            locked = next(candidate for candidate in request_lines if candidate.pk == line.pk)
            now = timezone.now()
            StoreRequisitionItem.objects.filter(pk=locked.pk).update(
                quantity_approved=Decimal("0.00"),
                storekeeper_comment=reason,
                rejection_stage="Store Keeper",
                rejection_reason=reason,
                rejected_at=now,
                rejected_by=request.user,
                updated_at=now,
            )
            locked.refresh_from_db()
        return Response(self.get_serializer(locked).data)

    def perform_destroy(self, instance):
        self._enforce_line_edit(instance.requisition)
        if instance.requisition.status not in (
            StoreRequisitionStatus.DRAFT,
            StoreRequisitionStatus.REJECTED,
        ):
            raise ValidationError("Only draft or rejected store request lines can be removed.")
        instance.delete()


class StockIssueViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockIssue.objects.select_related("requisition", "store", "issued_by", "received_by")
    serializer_class = StockIssueSerializer
    filterset_fields = ("requisition", "store", "issued_by", "received_by", "inventory_changes_applied")
    search_fields = ("issue_no", "requisition__requisition_no", "store__name", "received_by_name")
    ordering_fields = ("issue_no", "issue_date", "created_at")

    @action(detail=True, methods=["get"])
    def readiness(self, request, pk=None):
        return Response(self.get_object().posting_readiness())

    @action(detail=True, methods=["post"])
    def apply(self, request, pk=None):
        issue = self.get_object()
        enforce_readiness(issue.posting_readiness())
        try:
            issue.apply_inventory_changes()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(issue).data)

    @action(detail=True, methods=["post"])
    def acknowledge(self, request, pk=None):
        issue = self.get_object()
        if not issue.inventory_changes_applied:
            raise ValidationError("Stock must be issued before the department can acknowledge receipt.")
        received_by = None
        if request.data.get("received_by"):
            try:
                received_by = Employee.objects.get(pk=request.data["received_by"])
            except Employee.DoesNotExist:
                raise ValidationError({"received_by": "Selected employee was not found."})
        received_by_name = request.data.get("received_by_name", "").strip()
        if not received_by and not received_by_name:
            raise ValidationError("Select the receiving employee or enter the receiver's name.")
        issue.received_by = received_by
        issue.received_by_name = received_by_name or str(received_by)
        issue.save(update_fields=("received_by", "received_by_name", "updated_at"))
        return Response(self.get_serializer(issue).data)


class StockIssueItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockIssueItem.objects.select_related("issue", "requisition_item", "item", "unit")
    serializer_class = StockIssueItemSerializer
    filterset_fields = ("issue", "requisition_item", "item", "unit")
    search_fields = ("issue__issue_no", "item__name", "item__sku")
    ordering_fields = ("quantity", "base_quantity", "created_at")


class DepartmentConsumptionViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = DepartmentConsumption.objects.select_related(
        "department", "item", "stock_issue_item", "goods_receipt_item"
    )
    serializer_class = DepartmentConsumptionSerializer
    filterset_fields = ("department", "item", "consumed_on")
    search_fields = ("department__name", "item__name", "item__sku", "purpose")
    ordering_fields = ("consumed_on", "quantity", "unit_cost", "created_at")


class StoreReturnViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreReturn.objects.select_related("department", "store", "received_by")
    serializer_class = StoreReturnSerializer
    filterset_fields = ("department", "store", "received_by", "inventory_changes_applied")
    search_fields = ("return_no", "department__name", "store__name", "reason")
    ordering_fields = ("return_no", "return_date", "created_at")

    @action(detail=True, methods=["post"])
    def apply(self, request, pk=None):
        store_return = self.get_object()
        try:
            store_return.apply_inventory_changes()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(store_return).data)


class StoreReturnItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StoreReturnItem.objects.select_related("store_return", "item", "unit")
    serializer_class = StoreReturnItemSerializer
    filterset_fields = ("store_return", "item", "unit")
    search_fields = ("store_return__return_no", "item__name", "item__sku")
    ordering_fields = ("quantity", "base_quantity", "created_at")


class StockCountViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockCount.objects.select_related("store", "conducted_by", "approved_by")
    serializer_class = StockCountSerializer
    filterset_fields = ("store", "conducted_by", "approved_by", "status", "inventory_changes_applied")
    search_fields = ("count_no", "store__name", "note")
    ordering_fields = ("count_no", "count_date", "status", "created_at")

    @action(detail=True, methods=["post"])
    def populate(self, request, pk=None):
        stock_count = self.get_object()
        stock_count.populate_from_system_balances()
        return Response(self.get_serializer(stock_count).data)

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        stock_count = self.get_object()
        try:
            stock_count.submit()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(stock_count).data)

    @action(detail=True, methods=["post"], url_path="assign-store")
    def assign_store(self, request, pk=None):
        requisition = self.get_object()
        if not has_role(request.user, "System Administrator", "Store Keeper"):
            raise PermissionDenied("Only the Store Keeper can choose the destination store.")
        if requisition.status != StoreRequisitionStatus.SUBMITTED:
            raise ValidationError("The destination store can only be selected for a submitted Department request.")
        store_id = request.data.get("store")
        if not store_id:
            raise ValidationError({"store": "Select the destination store."})
        store = StoreLocation.objects.filter(pk=store_id, is_active=True).first()
        if not store:
            raise ValidationError({"store": "The selected destination store is not available."})
        if not has_role(request.user, "System Administrator") and not StoreKeeperAssignment.objects.filter(
            store=store, employee=getattr(request.user, "employee_profile", None), is_active=True
        ).exists():
            raise PermissionDenied("You can only route requests to a store assigned to you.")
        requisition.store = store
        requisition.save(update_fields=["store", "updated_at"])
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        stock_count = self.get_object()
        try:
            stock_count.approve(approved_by=getattr(request.user, "employee_profile", None))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(stock_count).data)

    @action(detail=True, methods=["post"])
    def cancel(self, request, pk=None):
        stock_count = self.get_object()
        try:
            stock_count.cancel()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(stock_count).data)

    @action(detail=True, methods=["post"])
    def apply(self, request, pk=None):
        stock_count = self.get_object()
        try:
            stock_count.apply_variances()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(stock_count).data)


class StockCountItemViewSet(CreatedByModelMixin, ModelViewSet):
    queryset = StockCountItem.objects.select_related("stock_count", "item")
    serializer_class = StockCountItemSerializer
    filterset_fields = ("stock_count", "item")
    search_fields = ("stock_count__count_no", "item__name", "item__sku")
    ordering_fields = ("system_quantity", "physical_quantity", "created_at")
    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment.submit()
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(adjustment).data)

    @action(detail=True, methods=["post"], url_path="assign-store")
    def assign_store(self, request, pk=None):
        requisition = self.get_object()
        if not has_role(request.user, "System Administrator", "Store Keeper"):
            raise PermissionDenied("Only the Store Keeper can choose the destination store.")
        if requisition.status != StoreRequisitionStatus.SUBMITTED:
            raise ValidationError("The destination store can only be selected for a submitted Department request.")
        store_id = request.data.get("store")
        if not store_id:
            raise ValidationError({"store": "Select the destination store."})
        store = StoreLocation.objects.filter(pk=store_id, is_active=True).first()
        if not store:
            raise ValidationError({"store": "The selected destination store is not available."})
        if not has_role(request.user, "System Administrator") and not StoreKeeperAssignment.objects.filter(
            store=store, employee=getattr(request.user, "employee_profile", None), is_active=True
        ).exists():
            raise PermissionDenied("You can only route requests to a store assigned to you.")
        requisition.store = store
        requisition.save(update_fields=["store", "updated_at"])
        return Response(self.get_serializer(requisition).data)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment.approve(approved_by=getattr(request.user, "employee_profile", None))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(adjustment).data)

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        adjustment = self.get_object()
        try:
            adjustment.reject(reason=request.data.get("reason", ""))
        except DjangoValidationError as error:
            raise_drf_validation_error(error)
        return Response(self.get_serializer(adjustment).data)
